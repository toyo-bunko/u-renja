import urllib.request
from bs4 import BeautifulSoup
import csv
from time import sleep
import pandas as pd
import json
import urllib.request
import os
from PIL import Image
import yaml
import requests

prefix = "https://static.toyobunko-lab.jp/u-renja"

with open("config.json") as f:
    config = json.load(f)

path = config["path"]

############

with open("../../static/iiif2/collection/top.json") as f:
    df2 = json.load(f)

manifests = df2["manifests"]

images = {}

for m in manifests:
    metadata = m["metadata"]

    num = -1
    identifier = ""

    for obj in metadata:
        if "Description" in obj["label"]:
            num = obj["value"][0].split("通番: ")[1]

            num = int(num)

        elif "identifier" in obj["label"]:
            identifier = obj["value"]

    if num not in images:
        images[num] = []

    images[num].append({
        "id": m["@id"],
        "identifier": identifier
    })

#############

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import openpyxl

hyperlink = Font(underline='single', color='0563C1')

wb = load_workbook(path)
ws = wb.active

# df_item = pd.read_excel(path, sheet_name="書名", header=None, index_col=None, engine="openpyxl")
# df_item2 = pd.read_excel(path, sheet_name="詳細", header=None, index_col=None, engine="openpyxl")

# r_count = len(df_item.index)
# c_count = len(df_item.columns)

r_count = ws.max_row
c_count = ws.max_column

map = {}

for i in range(0, c_count):
    # map[i] = df_item.iloc[0, i]
    map[i] = ws.cell(row=0+1, column=i+1).value

for j in range(1, r_count):

    print(str(j)+"/"+str(r_count))

    # id = df_item.iloc[j, 7]
    id = ws.cell(row=j+1, column=7+1).value

    # if pd.isnull(id):
    if id == "" or id == None:
        continue

    num = -1

    for i in map:
        label = map[i]
        # value = df_item.iloc[j, i]

        value = ws.cell(row=j+1, column=i+1).value
        # if not pd.isnull(value) and value != 0:
        if value != "" and value != None:

            value = str(value).strip()

            if label == "大正藏經典番號(1)":
                # df_item.iloc[j, i] = "[{}](https://static.toyobunko-lab.jp/taishozo/search/?keyword={})".format(value, value)
                ws.cell(row=j+1, column=i+1).value = '=HYPERLINK("https://static.toyobunko-lab.jp/taishozo/search/?keyword={}", "{}")'.format(value, value)
                ws.cell(row=j+1, column=i+1).font = hyperlink

            elif label == "通番":
                num = int(value)

            elif "画像フォルダ名" in label:
                # title = df_item.iloc[j, i-1]
                x = ws.cell(row=j+1, column=i+1+1).value
                y = ws.cell(row=j+1, column=i+1+2).value

                uuid = ws.cell(row=j+1, column=i+1).value + "_" + str(x).zfill(4) + "_" + str(y).zfill(4)

                title = ws.cell(row=j+1, column=i+1-1).value

                print(uuid, title)

                if num in images:
                    
                    arr = images[num]

                    # print(arr)

                    for a in arr:
                        # print("aaa", uuid, a["identifier"])
                        if uuid == a["identifier"]:
                            # df_item.iloc[j, i] = "[{}]({})".format(value, a["id"])
                            ws.cell(row=j+1, column=i+1-1).value = '=HYPERLINK("http://codh.rois.ac.jp/software/iiif-curation-viewer/demo/?manifest={}", "{}")'.format(a["id"], title)
                            ws.cell(row=j+1, column=i+1-1).font = hyperlink


opath = '../../static/metadata/data.xlsx'

'''
writer = pd.ExcelWriter('../../static/metadata/data.xlsx', engine="openpyxl")
df_item.to_excel(writer, sheet_name = '書名', index=False, header=False)
df_item2.to_excel(writer, sheet_name = '詳細', index=False, header=False)

#Excelファイルを保存
writer.save()
#Excelファイルを閉じる
writer.close()
'''

wb.save(opath)

